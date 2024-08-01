import openpyxl
from openpyxl.styles import Font
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.patches import Circle, RegularPolygon
from matplotlib.path import Path
from matplotlib.projections import register_projection
from matplotlib.projections.polar import PolarAxes
from matplotlib.spines import Spine
from matplotlib.transforms import Affine2D
import matplotlib.image as mpimg
#from PIL import Image
from scipy import stats
from matplotlib.colors import LinearSegmentedColormap
import os


def load_workbook(file_path):
    return openpyxl.load_workbook(file_path, keep_vba=True)


def calculate_score(value, thresholds):
    if value is None:
        return 1  # Default value for missing data

    # Extract only the numeric thresholds
    numeric_thresholds = [t for t in thresholds if isinstance(t, (int, float)) and t is not None]
    
    if not numeric_thresholds:
        return 1  # Default value if no valid thresholds are found

    # Determine if higher or lower values are better
    higher_is_better = numeric_thresholds[0] < numeric_thresholds[-1]

    if higher_is_better:
        sorted_thresholds = sorted(numeric_thresholds)
        for i, threshold in enumerate(sorted_thresholds):
            if value <= threshold:
                return i + 1
        return 5  # If value is GREATER than all thresholds
    else:
        sorted_thresholds = sorted(numeric_thresholds, reverse=True)
        for i, threshold in enumerate(sorted_thresholds):
            if value >= threshold:
                return i + 1
        return 5  # If value is lower than all thresholds

  
def process_sheet(sheet):
    total_score = 0
    total_weight = 0
    weighted_scores = {}
    section_scores = {}
   
    sections = group_parameters_by_section(sheet)
    print(f"Sections for {sheet.title}: {sections}")
    index = 1


    for row in sheet.iter_rows(min_row=2, max_col=10):
        if not row[1].value:  # Stop if we reach an empty row in column B
            break
        index+=1
        weight = row[1].value  # Column B
        param = row[2].value   # Column C
        value = row[3].value   # Column D
        thresholds = [cell.value for cell in row[4:9]]  # Columns E to I


       
        print(f"Processing row: weight={weight}, param={param}, value={value}")
       
        if weight is not None:
            score = calculate_score(value, thresholds)
            weighted_score = score * weight
            weighted_scores[param] = weighted_score
            # Write individual score to Output column (K)
            output_cell = sheet.cell(row=row[4].row, column=10)  # Column J
            output_cell.value = score






            # Add score to the appropriate section
            for section, params in sections.items():
                if param in params:
                    if section not in section_scores:
                        section_scores[section] = 0
                    section_scores[section] += weighted_score
                    break


            total_score += weighted_score
            total_weight += weight


    # Calculate final score out of 5
    final_score = total_score if total_weight > 0 else 0


    # Write final score at the bottom of the Output column
    last_row = index + 2
    final_score_cell = sheet.cell(row=last_row, column=10)  # Column M
    final_score_cell.value = final_score
    final_score_cell.font = Font(bold=True)
    sheet.cell(row=last_row, column=3, value="Final Score:").font = Font(bold=True)
   


    print(sections, section_scores)


    return final_score, weighted_scores, section_scores

def calculate_max_scores(sheet):
    max_scores = {}
    sections = group_parameters_by_section(sheet)

    for row in sheet.iter_rows(min_row=2, max_col=17):
        weight = row[1].value  # Column B
        param = row[2].value   # Column C

        if weight is not None and param is not None:
            max_param_score = 5 * weight  # Maximum score for each parameter is 5

            # Add the max score to the appropriate section
            for section, params in sections.items():
                if param in params:
                    if section not in max_scores:
                        max_scores[section] = 0
                    max_scores[section] += max_param_score
                    break

    return max_scores

def calculate_points_lost(section_scores, max_scores):
    points_lost = {}
    for section in section_scores:
        if section in max_scores:
            points_lost[section] = max_scores[section] - section_scores[section]
    return points_lost

def group_parameters_by_section(sheet):
    sections = {}
    current_section = None
    for row in sheet.iter_rows(min_row=2, max_col=6):
        if row[0].value:  # Column D
            current_section = row[0].value
            sections[current_section] = []
        if row[2].value and current_section:  # Column F (parameter)
            sections[current_section].append(row[2].value)
    print(f"Grouped sections: {sections}")  # Add this line for debugging
    return sections


def plot_normal_distribution(scores_with_names, plots):
    df = pd.DataFrame(scores_with_names, columns=['Score', 'Company'])
    
    # Sort the DataFrame by Score in descending order
    df = df.sort_values('Score', ascending=False)
    
    # Calculate mean and standard deviation
    mean = df['Score'].mean()
    std = df['Score'].std()

    # Create a range of x values
    x = np.linspace(mean - 3*std, mean + 3*std, 100)

    # Calculate the normal distribution
    y = stats.norm.pdf(x, mean, std)

    # Plot the histogram and the PDF
    plt.figure(figsize=(12, 8))  # Increased figure size to accommodate legend

    plt.plot(x, y, color='green', marker=',', linestyle='dotted', linewidth=2, markersize=10)

    y_scatter = stats.norm.pdf(df['Score'], mean, std)
    
    # Generate a color map
    companies = df['Company'].unique()
    num_companies = len(companies)
    color_map = plt.get_cmap('tab20')
    colors = {company: color_map(i % color_map.N) for i, company in enumerate(companies)}

    # Scatter plot with different colors for each company
    for company in companies:
        company_data = df[df['Company'] == company]
        plt.scatter(company_data['Score'], stats.norm.pdf(company_data['Score'], mean, std), 
                    alpha=1.0, color=colors[company], label=company)

    plt.title('Normal Distribution of Scores')
    plt.xlabel('Score')
    plt.ylabel('Density')
    
    # Add mean and std dev to the plot
    plt.axvline(mean, color='deepskyblue', linestyle='dashed', linewidth=2)
    plt.axvline(mean+std, color='red', linestyle='dotted', linewidth=2)
    plt.axvline(mean-std, color='red', linestyle='dotted', linewidth=2)
    plt.text(mean*1.1, plt.ylim()[1]*0.9, f'Mean: {mean:.2f}', color='blue')
    plt.text(mean*1.1, plt.ylim()[1]*0.8, f'SD: {std:.2f}', color='red')
    plt.text(mean + std, plt.ylim()[1]*0.2, 'A Rated', color='blue', ha='left', va='bottom')
    plt.text(mean - std, plt.ylim()[1]*0.2, 'Unrated', color='blue', ha='right', va='bottom')
    plt.text(mean, plt.ylim()[1]*0.2, 'B Rated', color='blue', ha='center', va='bottom')

    # Add legend
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)

    plt.tight_layout()
    plot_path = os.path.abspath('myplot.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'A1'
    plots.add_image(img)
    plt.close()

def plot_pie_chart(sheet_name, weighted_scores, sheet):
    plt.figure(figsize=(12, 8))
    labels = list(weighted_scores.keys())
    sizes = list(weighted_scores.values())
    
    # Calculate total score and points lost
    total_score = sum(sizes)
    points_lost = 5.0 - total_score

    # Add points lost to the data
    sizes.append(points_lost)
    labels.append("Points Lost")

    # Sort data by size
    sorted_data = sorted(zip(sizes, labels), reverse=True)
    sizes, labels = zip(*sorted_data)
    
    # Group small slices
    threshold = 0.07  # Adjust as needed
    other_sum = sum(size for size in sizes if size < threshold)

    sizes = [size for size in sizes if size >= threshold] + [other_sum]
    labels = [label for size, label in zip(sorted_data, labels) if size[0] >= threshold] + ["Other"] 
    
    # Create a custom color map with more unique colors
    n_colors = len(sizes)
    colors = list(plt.cm.tab20(np.arange(n_colors) / n_colors))

    # Find the index of "Points Lost" and set its color to black
    points_lost_index = labels.index("Points Lost") if "Points Lost" in labels else -1
    if points_lost_index != -1:
        colors[points_lost_index] = 'black'
    
    # Create pie chart
    wedges, texts, autotexts = plt.pie(sizes, labels=None, autopct='%1.1f%%', 
                                       colors=colors, startangle=90, 
                                       pctdistance=0.85, labeldistance=1.05)
    
    # Add a legend
    plt.legend(wedges, labels, title="Categories", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    plt.title(f'Weighted Scores Distribution - {sheet_name}')
    plt.axis('equal')
    
    # Adjust layout to prevent clipping of labels
    plt.tight_layout()
    
    plot_path = os.path.abspath(f'pie_chart_{sheet_name}.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'N1'  # This will place the pie chart in column N, row 1
    sheet.add_image(img)
    plt.close()
    #os.remove(plot_path)  # Remove the temporary image file




#radar chart design from web - matplotlib (radar chart)
def radar_factory(num_vars, frame='circle'):
    """
    Create a radar chart with `num_vars` Axes.

    This function creates a RadarAxes projection and registers it.

    Parameters
    ----------
    num_vars : int
        Number of variables for radar chart.
    frame : {'circle', 'polygon'}
        Shape of frame surrounding Axes.

    """
    # calculate evenly-spaced axis angles
    theta = np.linspace(0, 2*np.pi, num_vars, endpoint=False)

    class RadarTransform(PolarAxes.PolarTransform):

        def transform_path_non_affine(self, path):
            # Paths with non-unit interpolation steps correspond to gridlines,
            # in which case we force interpolation (to defeat PolarTransform's
            # autoconversion to circular arcs).
            if path._interpolation_steps > 1:
                path = path.interpolated(num_vars)
            return Path(self.transform(path.vertices), path.codes)

    class RadarAxes(PolarAxes):

        name = 'radar'
        PolarTransform = RadarTransform

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            # rotate plot such that the first axis is at the top
            self.set_theta_zero_location('N')

        def fill(self, *args, closed=True, **kwargs):
            """Override fill so that line is closed by default"""
            return super().fill(closed=closed, *args, **kwargs)

        def plot(self, *args, **kwargs):
            """Override plot so that line is closed by default"""
            lines = super().plot(*args, **kwargs)
            for line in lines:
                self._close_line(line)

        def _close_line(self, line):
            x, y = line.get_data()
            # FIXME: markers at x[0], y[0] get doubled-up
            if x[0] != x[-1]:
                x = np.append(x, x[0])
                y = np.append(y, y[0])
                line.set_data(x, y)

        def set_varlabels(self, labels):
            self.set_thetagrids(np.degrees(theta), labels)

        def _gen_axes_patch(self):
            # The Axes patch must be centered at (0.5, 0.5) and of radius 0.5
            # in axes coordinates.
            if frame == 'circle':
                return Circle((0.5, 0.5), 0.5)
            elif frame == 'polygon':
                return RegularPolygon((0.5, 0.5), num_vars,
                                      radius=.5, edgecolor="k")
            else:
                raise ValueError("Unknown value for 'frame': %s" % frame)

        def _gen_axes_spines(self):
            if frame == 'circle':
                return super()._gen_axes_spines()
            elif frame == 'polygon':
                # spine_type must be 'left'/'right'/'top'/'bottom'/'circle'.
                spine = Spine(axes=self,
                              spine_type='circle',
                              path=Path.unit_regular_polygon(num_vars))
                # unit_regular_polygon gives a polygon of radius 1 centered at
                # (0, 0) but we want a polygon of radius 0.5 centered at (0.5,
                # 0.5) in axes coordinates.
                spine.set_transform(Affine2D().scale(.5).translate(.5, .5)
                                    + self.transAxes)
                return {'polar': spine}
            else:
                raise ValueError("Unknown value for 'frame': %s" % frame)

    register_projection(RadarAxes)
    return theta


def plot_spider(all_section_scores, plots):

    categories = list(all_section_scores.keys())
    values = [all_section_scores[cat] for cat in categories]

    # Number of variables
    num_vars = len(categories)


    # Compute angle for each variable
    theta = np.linspace(0, 2*np.pi, num_vars, endpoint=False)

    # Create figure with two subplots: one for the spider chart, one for the table
    fig, (ax, ax_table) = plt.subplots(2, 1, figsize=(12, 12), 
                                       gridspec_kw={'height_ratios': [3, 1]})
    
    # Set the first subplot (ax) to polar projection for the spider chart
    ax = plt.subplot(2, 1, 1, projection='polar')

    # Close the polygon by appending the first value to the end
    values_closed = np.concatenate((values, [values[0]]))
    theta_closed = np.concatenate((theta, [theta[0]]))

    # Plot data on the spider chart
    ax.plot(theta_closed, values_closed, 'o-', linewidth=2, color='blue')
    ax.fill(theta_closed, values_closed, alpha=0.25, color='skyblue')

    # Set category labels
    ax.set_xticks(theta)
    ax.set_xticklabels(categories)

    # Set y-ticks
    ax.set_yticks([0.2, 0.4, 0.6, 0.8])
    ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8'])
    ax.set_ylim(0, 1)

    # Add legend
    ax.legend(['Weighted Scores', 'Score Area'], loc='upper right', bbox_to_anchor=(1.3, 1.1))

    # Add title
    ax.set_title(f'Weighted Scores by Section', y=1.1, fontweight='bold')

    # Add table with values
    ax_table.axis('off')
    table_data = [[cat, f'{val:.2f}'] for cat, val in zip(categories, values)]
    table = ax_table.table(cellText=table_data, colLabels=['Category', 'Score'], 
                           loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.5)

    # Save the plot
    plot_path = os.path.abspath(f'spider_chart.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'O1'  # Adjust this to place the chart where you want
    plots.add_image(img)
    plt.close()

    return plot_path  # Return the path for potential use in the main function
    
    
def plot_spider_chart(section_scores, sheet_name, sheet):
    categories = list(section_scores.keys())
    values = [section_scores[cat] for cat in categories]

    # Number of variables
    num_vars = len(categories)

    # Check if there are any variables to plot
    if num_vars == 0:
        print(f"No data to plot for {sheet_name}")
        return

    # Compute angle for each variable
    theta = np.linspace(0, 2*np.pi, num_vars, endpoint=False)
    
    # Create figure with two subplots: one for the spider chart, one for the table
    fig, (ax, ax_table) = plt.subplots(2, 1, figsize=(12, 12), 
                                       gridspec_kw={'height_ratios': [3, 1]})
    
    fig, (ax, ax_table) = plt.subplots(2, 1, figsize=(14, 14), 
                                   gridspec_kw={'height_ratios': [3, 1]})
    

    # Set the first subplot (ax) to polar projection for the spider chart
    ax = plt.subplot(2, 1, 1, projection='polar')

    # Close the polygon by appending the first value to the end
    values_closed = np.concatenate((values, [values[0]]))
    theta_closed = np.concatenate((theta, [theta[0]]))

    # Plot data on the spider chart
    ax.plot(theta_closed, values_closed, 'o-', linewidth=2, color='blue')
    ax.fill(theta_closed, values_closed, alpha=0.25, color='skyblue')

    # Set category labels
    ax.set_xticks(theta)
    ax.set_xticklabels(categories)

    # Set y-ticks
    ax.set_yticks([0.2, 0.4, 0.6, 0.8])
    ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8'])
    ax.set_ylim(0, 1)

    # Add legend
    ax.legend(['Weighted Scores', 'Score Area'], loc='upper right', bbox_to_anchor=(1.3, 1.1))

    # Add title
    ax.set_title(f'Weighted Scores by Section - {sheet_name}', y=1.1, fontweight='bold')


    # Add table with values
    ax_table.axis('off')
    table_data = [[cat, f'{val:.2f}'] for cat, val in zip(categories, values)]
    table = ax_table.table(cellText=table_data, colLabels=['Category', 'Score'], 
                           loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.5)

    # Save the plot
    plot_path = os.path.abspath(f'spider_chart_{sheet_name}.png')
    plt.savefig(plot_path, dpi=60, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'AA1'  # Adjust this to place the chart where you want
    sheet.add_image(img)
    plt.close()

    return plot_path  # Return the path for potential use in the main function

def plot_points_lost_histogram(points_lost, sheet_name, sheet):
    sections = list(points_lost.keys())
    lost = list(points_lost.values())

    plt.figure(figsize=(12, 6))
    bars = plt.bar(sections, lost)

    plt.title(f'Points Lost by Section - {sheet_name}')
    plt.xlabel('Sections')
    plt.ylabel('Points Lost')
    plt.xticks(rotation=45, ha='right')

    # Add value labels on top of each bar
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height,
                 f'{height:.2f}',
                 ha='center', va='bottom')

    plt.tight_layout()
    
    plot_path = os.path.abspath(f'points_lost_histogram_{sheet_name}.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'N32'  # Adjust this to place the chart where you want
    sheet.add_image(img)
    plt.close()

def clear_plots_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
    for img in sheet._images:
        sheet._images.remove(img)

def clear_plots(sheet):
    for img in sheet._images:
        sheet._images.remove(img)

def main():
    file_path = 'python_template2.xlsm'  # Replace with your actual file path
    workbook = load_workbook(file_path)
    plots = workbook['Plots']
    #clear_plots_sheet(plots)
    
    all_section_scores = {}
    scores_with_names = []

    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Plots':
            continue
        print(f"\nProcessing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        clear_plots(sheet)
        final_score, weighted_scores, section_scores = process_sheet(sheet)
        scores_with_names.append((final_score, sheet_name))
        all_section_scores |= section_scores
        
        max_scores = calculate_max_scores(sheet)
        # Calculate points lost
        points_lost = calculate_points_lost(section_scores, max_scores)

        #clear_plots(sheet)
        # Plot histogram of points lost
        plot_points_lost_histogram(points_lost, sheet_name, sheet)



        print(f"Weighted scores: {weighted_scores}")
        print(f"Section scores: {section_scores}")
        

        if section_scores:
            plot_pie_chart(sheet_name, weighted_scores, sheet)
            plot_spider_chart(section_scores, sheet_name, sheet)
        else:
            print(f"No data to plot for {sheet_name}")
        
        print(f"Final Score for sheet '{sheet_name}': {final_score:.2f}/5")
    
    plot_normal_distribution(scores_with_names, plots)
    plot_spider(all_section_scores, plots)
    
    workbook.save(file_path)

if __name__ == "__main__":
    main()

